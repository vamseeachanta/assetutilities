#!/usr/bin/env python3
"""
Simple test for parallel Excel utilities processing.
Tests the basic functionality without complex path resolution.
"""
import os
import sys
import time
import yaml
from pathlib import Path

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent.parent.parent))

from assetutilities.modules.excel_utilities.excel_utilities_parallel import ExcelUtilitiesParallel


def test_parallel_excel():
    """Simple test of parallel Excel processing."""
    
    # Setup paths - use simple relative paths from test directory
    test_dir = Path(__file__).parent
    os.chdir(test_dir)  # Change to test directory
    
    # Create simple test configuration
    cfg = {
        'meta': {
            'basename': 'excel_utilities',
            'library': 'assetutilities'
        },
        'basename': 'excel_utilities',
        'task': 'csv_copy_to_excel',
        'parallel_processing': {
            'enabled': True,
            'max_workers': 'auto',
            'task_type': 'io_bound'
        },
        'data': {
            'input_format': 'csv',
            'groups': [
                {
                    'input': {
                        'filename': 'input_data/inputs.csv'
                    },
                    'target': {
                        'filename': 'input_data/target.xlsx',
                        'sheet_name': 'inputs'
                    }
                },
                {
                    'input': {
                        'filename': 'input_data/results.csv'
                    },
                    'target': {
                        'filename': 'input_data/target.xlsx',
                        'sheet_name': 'results'
                    }
                }
            ]
        },
        'Analysis': {
            'analysis_root_folder': '.',  # Current directory (test directory)
            'result_folder': './output'
        },
        'default': {
            'log_level': 'DEBUG',
            'config': {
                'overwrite': {
                    'output': True
                }
            }
        }
    }
    
    # Create output directory
    output_dir = Path('output')
    output_dir.mkdir(exist_ok=True)
    
    # Ensure input files exist
    input_dir = Path('input_data')
    input_dir.mkdir(exist_ok=True)
    
    # Create test CSV files if they don't exist
    inputs_csv = input_dir / 'inputs.csv'
    if not inputs_csv.exists():
        import pandas as pd
        df = pd.DataFrame({
            'ID': [1, 2, 3],
            'Name': ['Item1', 'Item2', 'Item3'],
            'Value': [10, 20, 30]
        })
        df.to_csv(inputs_csv, index=False)
        print(f"Created: {inputs_csv}")
    
    results_csv = input_dir / 'results.csv'
    if not results_csv.exists():
        import pandas as pd
        df = pd.DataFrame({
            'TestID': [1, 2],
            'Status': ['Pass', 'Fail'],
            'Score': [95, 45]
        })
        df.to_csv(results_csv, index=False)
        print(f"Created: {results_csv}")
    
    # Create target Excel file if it doesn't exist
    target_xlsx = input_dir / 'target.xlsx'
    if not target_xlsx.exists():
        from openpyxl import Workbook
        wb = Workbook()
        wb.save(target_xlsx)
        print(f"Created: {target_xlsx}")
    
    print("\n" + "=" * 60)
    print("TESTING PARALLEL EXCEL PROCESSING")
    print("=" * 60)
    
    # Initialize Excel utilities
    excel_util = ExcelUtilitiesParallel()
    
    # Test parallel processing
    print("\nRunning parallel processing...")
    start_time = time.time()
    
    try:
        result = excel_util.excel_utility_router(cfg)
        elapsed = time.time() - start_time
        
        print(f"\nCompleted in {elapsed:.3f} seconds")
        
        # Check results
        if 'processing_results' in result:
            results = result['processing_results']
            successful = sum(1 for r in results if r['success'])
            total = len(results)
            
            print(f"\nResults:")
            print(f"  Successful: {successful}/{total}")
            
            for i, res in enumerate(results, 1):
                status = "OK" if res['success'] else "FAIL"
                print(f"  Group {i}: {status} - {res['message']}")
            
            if successful == total:
                print(f"\nSUCCESS: All {total} groups processed successfully!")
                
                # Verify the Excel file was created/updated
                if target_xlsx.exists():
                    from openpyxl import load_workbook
                    wb = load_workbook(target_xlsx)
                    sheets = wb.sheetnames
                    print(f"\nTarget Excel file sheets: {sheets}")
                    
                    for sheet_name in ['inputs', 'results']:
                        if sheet_name in sheets:
                            ws = wb[sheet_name]
                            row_count = ws.max_row
                            col_count = ws.max_column
                            print(f"  {sheet_name}: {row_count} rows, {col_count} columns")
                    wb.close()
                    
                return True
            else:
                print(f"\nFAILED: Only {successful}/{total} groups succeeded")
                return False
        else:
            print("\nERROR: No processing results found")
            return False
            
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_sequential_vs_parallel():
    """Compare sequential and parallel processing."""
    
    test_dir = Path(__file__).parent
    os.chdir(test_dir)
    
    # Simple configuration
    base_cfg = {
        'meta': {'basename': 'excel_utilities'},
        'task': 'csv_copy_to_excel',
        'data': {
            'groups': [
                {
                    'input': {'filename': 'input_data/inputs.csv'},
                    'target': {'filename': 'input_data/target1.xlsx', 'sheet_name': 'data'}
                },
                {
                    'input': {'filename': 'input_data/results.csv'},
                    'target': {'filename': 'input_data/target2.xlsx', 'sheet_name': 'data'}
                }
            ]
        },
        'Analysis': {'analysis_root_folder': '.'}
    }
    
    excel_util = ExcelUtilitiesParallel()
    
    # Sequential
    seq_cfg = base_cfg.copy()
    seq_cfg['parallel_processing'] = {'enabled': False}
    
    print("\nSequential processing...")
    start = time.time()
    seq_result = excel_util.csv_copy_to_excel_sequential(seq_cfg)
    seq_time = time.time() - start
    
    # Parallel
    par_cfg = base_cfg.copy()
    par_cfg['parallel_processing'] = {'enabled': True}
    
    print("Parallel processing...")
    start = time.time()
    par_result = excel_util.csv_copy_to_excel_parallel(par_cfg)
    par_time = time.time() - start
    
    print(f"\nTiming:")
    print(f"  Sequential: {seq_time:.3f}s")
    print(f"  Parallel:   {par_time:.3f}s")
    
    if par_time < seq_time:
        print(f"  Speedup:    {seq_time/par_time:.2f}x")


if __name__ == "__main__":
    print("Starting Excel Utilities Parallel Test")
    print("Current directory:", os.getcwd())
    
    # Run tests
    success = test_parallel_excel()
    
    if success:
        print("\n" + "=" * 60)
        test_sequential_vs_parallel()
    
    sys.exit(0 if success else 1)