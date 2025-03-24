import os
import logging
import excel2img
import pandas as pd
from openpyxl import load_workbook
import shutil


import win32api

from xlsxwriter.workbook import Workbook
from xlsxwriter.utility import xl_rowcol_to_cell

from assetutilities.common.data import ReadFromExcel
from assetutilities.calculations.polynomial import Polynomial
from assetutilities.common.utilities import is_file_valid_func


rfe = ReadFromExcel()
_polynomial = Polynomial()


class ExcelUtilities:

    def __init__(self) -> None:
        pass

    def router(self, cfg):
        if cfg["task"] == "cross_reference_values_from_closed_workbooks":
            self.cross_reference_values_from_closed_workbooks_xlsxwriter(cfg)
        if cfg["task"] == "custom_calculation":
            data = self.get_data(cfg)

        return cfg

    def excel_utility_router(self, cfg):
        if cfg["task"] == "cross_reference_values_from_closed_workbooks":
            self.cross_reference_values_from_closed_workbooks_xlsxwriter(cfg)
        if cfg["task"] == "custom_calculation":
            data = self.get_data(cfg)
        if cfg["task"] == "excel_to_image":
            data = self.excel_to_image(cfg)
        if cfg['task'] == 'csv_copy_to_excel':
            data = self.csv_copy_to_excel(cfg)

        return cfg

    def cross_reference_values_from_closed_workbooks_xlsxwriter(self, cfg):

        workbook = Workbook(
            cfg["files"]["target"]["path"] + cfg["files"]["target"]["workbook"]
        )
        worksheet = workbook.add_worksheet(cfg["files"]["target"]["worksheet"])
        worksheet.write_formula("A2", "=VLOOKUP(B3,'Sheet2'!$B$2:$R$2199,17,FALSE)")

        cell_range = xl_rowcol_to_cell(4, 1)
        cell_formula = (
            cfg["files"]["source"]["path"]
            + "["
            + cfg["files"]["source"]["workbook"]
            + "]"
            + cfg["files"]["source"]["worksheet"]
            + "!"
            + "BH5"
        )
        # TODO did not work with abstraction
        # worksheet.write_formula(cell_range, cell_formula)

        worksheet.write_formula("A3", "[S01-Dyn-FC180-2.5mHs.xlsx]EditingTable'!BH5")

        # worksheet.write_formula(
        #     'A3', "'[S01-Dyn-FC180-2.5mHs.xlsx]EditingTable'!BH5")
        # worksheet.write_formula('A2',"=A20")

        # workbook.add_worksheet('Sheet2')

        workbook.close()

    def get_data(self, cfg):
        pass

    def excel_to_image(self, cfg):

        for file in cfg["files"]:
            io = file["io"]
            for sheetname in file["sheet_name"]:
                for array_range in file["range"]:
                    cell_range = array_range[0] + ":" + array_range[1]
                    sheet_range = sheetname + "!" + cell_range
                    output_basename = sheetname + "_" + array_range[0] + array_range[1]
                    for ext in file["output_extension"]:
                        output_filename = output_basename + "." + ext
                        if file["output_dir"] is not None and os.path.isdir(
                            file["output_dir"]
                        ):
                            output_filename = os.path.join(
                                file["output_dir"], output_filename
                            )
                        else:
                            output_filename = os.path.join(
                                cfg["Analysis"]["result_folder"], output_filename
                            )
                        excel2img.export_img(io, output_filename, "", sheet_range)

    def csv_copy_to_excel(self, cfg):

        groups = cfg['data']['groups']
        for group in groups:
            analysis_root_folder = cfg['Analysis']['analysis_root_folder']

            template_file = group['target']['template']
            is_file_valid, template_file = is_file_valid_func(template_file, analysis_root_folder)

            target_file = group['target']['filename']
            is_file_valid, target_file = is_file_valid_func(target_file, analysis_root_folder)
            if not is_file_valid:
                target_file = os.path.join(analysis_root_folder, target_file)

            # Copy the template file to the target file
            shutil.copy(template_file, target_file)
            logging.debug(f"Template file copied to: {target_file}")

            csvs = group['csvs']

            # Load the target Excel file
            for csv in csvs:
                inputs_csv = csv['input']['filename']
                is_file_valid, inputs_csv = is_file_valid_func(inputs_csv, analysis_root_folder)
                df = pd.read_csv(inputs_csv)

                wb = load_workbook(target_file, data_only=False)  # Keep formulas intact
                sheet_name = csv['target']['sheet_name']
                worksheet = wb[sheet_name]
                for row in worksheet['A1:BZ1000']:
                    for cell in row:
                        cell.value = None
                wb.save(target_file)

                with pd.ExcelWriter(target_file, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0, startcol=0)

                logging.info(f"CSV data, {worksheet} copied to Excel file: {target_file}")