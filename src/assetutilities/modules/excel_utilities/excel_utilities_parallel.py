"""
Enhanced Excel Utilities with parallel processing support.
Utilizes parallel processing architecture for improved performance.
"""
import os
import shutil
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Dict, Any, List, Tuple

import excel2img
import pandas as pd
from loguru import logger
from openpyxl import load_workbook
from xlsxwriter.workbook import Workbook

from assetutilities.calculations.polynomial import Polynomial
from assetutilities.common.data import ReadFromExcel
from assetutilities.common.utilities import is_file_valid_func

rfe = ReadFromExcel()
_polynomial = Polynomial()


class ParallelProcessingConfig:
    """Configuration for parallel processing."""
    
    DEFAULT_WORKERS_IO_BOUND = 8  # Optimized for I/O-bound Excel operations
    
    @staticmethod
    def get_optimal_workers(num_tasks: int, task_type: str = 'io_bound') -> int:
        """
        Determine optimal number of workers based on task count and type.
        
        Args:
            num_tasks: Number of tasks to process
            task_type: Type of task ('io_bound' or 'cpu_bound')
            
        Returns:
            Optimal number of workers
        """
        import multiprocessing
        cpu_cores = multiprocessing.cpu_count()
        
        if task_type == 'io_bound':
            # For I/O-bound Excel operations, use limited workers to avoid disk contention
            base_workers = ParallelProcessingConfig.DEFAULT_WORKERS_IO_BOUND
        else:
            base_workers = cpu_cores
        
        # Don't use more workers than tasks
        workers = min(base_workers, num_tasks) if num_tasks > 0 else base_workers
        
        # Ensure at least 1 worker
        return max(1, workers)


class ExcelUtilitiesParallel:
    """Excel utilities with parallel processing capabilities."""
    
    def __init__(self) -> None:
        self.parallel_config = ParallelProcessingConfig()
    
    def router(self, cfg):
        if cfg["task"] == "cross_reference_values_from_closed_workbooks":
            self.cross_reference_values_from_closed_workbooks_xlsxwriter(cfg)
        if cfg["task"] == "custom_calculation":
            self.get_data(cfg)
        return cfg
    
    def excel_utility_router(self, cfg):
        if cfg["task"] == "cross_reference_values_from_closed_workbooks":
            self.cross_reference_values_from_closed_workbooks_xlsxwriter(cfg)
        if cfg["task"] == "custom_calculation":
            self.get_data(cfg)
        if cfg["task"] == "excel_to_image":
            self.excel_to_image(cfg)
        if cfg["task"] == "csv_copy_to_excel":
            # Use parallel processing for csv_copy_to_excel
            self.csv_copy_to_excel_parallel(cfg)
        return cfg
    
    def cross_reference_values_from_closed_workbooks_xlsxwriter(self, cfg):
        workbook = Workbook(
            cfg["files"]["target"]["path"] + cfg["files"]["target"]["workbook"]
        )
        worksheet = workbook.add_worksheet(cfg["files"]["target"]["worksheet"])
        worksheet.write_formula("A2", "=VLOOKUP(B3,'Sheet2'!$B$2:$R$2199,17,FALSE)")
        worksheet.write_formula("A3", "[S01-Dyn-FC180-2.5mHs.xlsx]EditingTable'!BH5")
        workbook.close()
    
    def get_data(self, cfg):
        pass
    
    def excel_to_image(self, cfg):
        for file in cfg["files"]:
            io = file["io"]
            io_basename = os.path.splitext(os.path.basename(io))[0]
            analysis_root_folder = cfg["Analysis"]["analysis_root_folder"]
            is_file_valid, io = is_file_valid_func(io, analysis_root_folder)
            for sheetname in file["sheet_name"]:
                for array_range in file["range"]:
                    cell_range = array_range[0] + ":" + array_range[1]
                    sheet_range = sheetname + "!" + cell_range
                    output_basename = (
                        io_basename
                        + "_"
                        + sheetname
                        + "_"
                        + array_range[0]
                        + array_range[1]
                    )
                    for ext in file["output_extension"]:
                        output_filename = output_basename + "." + ext
                        if file["output_dir"] is not None and os.path.isdir(
                            file["output_dir"]
                        ):
                            output_filename = os.path.join(
                                file["output_dir"], output_filename
                            )
                        else:
                            output_filename = os.path.join(
                                cfg["Analysis"]["result_folder"], output_filename
                            )
                        excel2img.export_img(io, output_filename, "", sheet_range)
    
    def _process_single_group(self, group: Dict[str, Any], cfg: Dict[str, Any]) -> Tuple[bool, str]:
        """
        Process a single group of CSV to Excel operations.
        
        Args:
            group: Group configuration containing input and target information
            cfg: Main configuration dictionary
            
        Returns:
            Tuple of (success, message)
        """
        try:
            # Get analysis root folder or use current directory
            analysis_root_folder = cfg.get("Analysis", {}).get("analysis_root_folder", "")
            if not analysis_root_folder:
                analysis_root_folder = os.getcwd()
            
            # Get target file information
            target_info = group.get("target", {})
            target_file = target_info.get("filename", "")
            
            if not target_file:
                return False, "No target filename specified"
            
            is_file_valid, target_file = is_file_valid_func(target_file, analysis_root_folder)
            if not is_file_valid:
                target_file = os.path.join(analysis_root_folder, target_file)
            
            # Process input CSV
            input_info = group.get("input", {})
            input_csv = input_info.get("filename", "")
            
            if not input_csv:
                return False, f"No input CSV specified for {target_file}"
            
            is_file_valid, input_csv = is_file_valid_func(input_csv, analysis_root_folder)
            
            if not is_file_valid:
                logger.debug(f"File {input_csv} not found. Skipping {target_file}")
                if os.path.exists(target_file):
                    logger.info(f"Removing {target_file} as it is not updated")
                    os.remove(target_file)
                return False, f"Input file {input_csv} not found"
            
            # Read CSV data
            df = pd.read_csv(input_csv)
            
            # Load or create workbook
            sheet_name = target_info.get("sheet_name", "Sheet1")
            
            if os.path.exists(target_file):
                # Load existing workbook
                wb = load_workbook(target_file, data_only=False)
                wb_sheetnames = wb.sheetnames
                
                if sheet_name not in wb_sheetnames:
                    logger.info(f"Sheet {sheet_name} does not exist in {target_file}. Creating it.")
                    wb.create_sheet(sheet_name)
                
                # Clear existing data in the sheet
                worksheet = wb[sheet_name]
                for row in worksheet["A1:BZ1000"]:
                    for cell in row:
                        cell.value = None
                wb.save(target_file)
            else:
                # Create new workbook if file doesn't exist
                logger.info(f"Creating new workbook: {target_file}")
                wb = Workbook(target_file)
                ws = wb.add_worksheet(sheet_name)
                wb.close()
            
            # Write CSV data to Excel
            with pd.ExcelWriter(
                target_file,
                mode="a" if os.path.exists(target_file) else "w",
                engine="openpyxl",
                if_sheet_exists="overlay" if os.path.exists(target_file) else None,
            ) as writer:
                df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                    startrow=0,
                    startcol=0,
                )
            
            logger.info(f"Successfully copied {input_csv} to {target_file}:{sheet_name}")
            return True, f"Processed {input_csv} -> {target_file}:{sheet_name}"
            
        except Exception as e:
            error_msg = f"Error processing group: {str(e)}"
            logger.error(error_msg)
            return False, error_msg
    
    def csv_copy_to_excel_parallel(self, cfg: Dict[str, Any]) -> Dict[str, Any]:
        """
        Copy CSV data to Excel files using parallel processing for groups.
        
        Args:
            cfg: Configuration dictionary containing data groups and settings
            
        Returns:
            Updated configuration with processing results
        """
        groups = cfg.get("data", {}).get("groups", [])
        
        if not groups:
            logger.warning("No groups found in configuration")
            return cfg
        
        # Check if parallel processing is enabled
        parallel_config = cfg.get("parallel_processing", {})
        enabled = parallel_config.get("enabled", True)
        
        if not enabled or len(groups) <= 1:
            # Fall back to sequential processing
            logger.info("Using sequential processing")
            return self.csv_copy_to_excel_sequential(cfg)
        
        # Determine optimal number of workers
        num_workers = self.parallel_config.get_optimal_workers(
            num_tasks=len(groups),
            task_type='io_bound'
        )
        
        logger.info(f"Processing {len(groups)} groups in parallel with {num_workers} workers")
        
        # Process groups in parallel
        results = []
        with ThreadPoolExecutor(max_workers=num_workers) as executor:
            # Submit all tasks
            future_to_group = {
                executor.submit(self._process_single_group, group, cfg): group
                for group in groups
            }
            
            # Collect results as they complete
            for future in as_completed(future_to_group):
                group = future_to_group[future]
                try:
                    success, message = future.result()
                    results.append({
                        'group': group,
                        'success': success,
                        'message': message
                    })
                    if success:
                        logger.info(f"✓ {message}")
                    else:
                        logger.warning(f"✗ {message}")
                except Exception as exc:
                    logger.error(f"Group generated an exception: {exc}")
                    results.append({
                        'group': group,
                        'success': False,
                        'message': str(exc)
                    })
        
        # Store results in configuration
        cfg['processing_results'] = results
        
        # Log summary
        successful = sum(1 for r in results if r['success'])
        logger.info(f"Parallel processing complete: {successful}/{len(groups)} groups processed successfully")
        
        return cfg
    
    def csv_copy_to_excel_sequential(self, cfg: Dict[str, Any]) -> Dict[str, Any]:
        """
        Fallback sequential processing for backward compatibility.
        
        Args:
            cfg: Configuration dictionary
            
        Returns:
            Updated configuration
        """
        groups = cfg.get("data", {}).get("groups", [])
        results = []
        
        for group in groups:
            success, message = self._process_single_group(group, cfg)
            results.append({
                'group': group,
                'success': success,
                'message': message
            })
            if success:
                logger.info(f"✓ {message}")
            else:
                logger.warning(f"✗ {message}")
        
        cfg['processing_results'] = results
        return cfg


# Backward compatibility alias
ExcelUtilities = ExcelUtilitiesParallel