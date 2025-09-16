"""
Enhanced Excel Utilities with parallel processing for different configuration structures.
Handles both simple group-based configs and complex CSV-list-based configs.
"""
import os
import shutil
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Dict, Any, List, Tuple
import time

import pandas as pd
from loguru import logger
from openpyxl import load_workbook
from xlsxwriter.workbook import Workbook

from assetutilities.common.utilities import is_file_valid_func


class ParallelProcessingConfig:
    """Configuration for parallel processing."""
    
    DEFAULT_WORKERS_IO_BOUND = 8  # Optimized for I/O-bound Excel operations
    
    @staticmethod
    def get_optimal_workers(num_tasks: int, task_type: str = 'io_bound') -> int:
        """Determine optimal number of workers based on task count and type."""
        import multiprocessing
        cpu_cores = multiprocessing.cpu_count()
        
        if task_type == 'io_bound':
            base_workers = ParallelProcessingConfig.DEFAULT_WORKERS_IO_BOUND
        else:
            base_workers = cpu_cores
        
        workers = min(base_workers, num_tasks) if num_tasks > 0 else base_workers
        return max(1, workers)


class ExcelUtilitiesEnhanced:
    """Enhanced Excel utilities with parallel processing for complex configurations."""
    
    def __init__(self) -> None:
        self.parallel_config = ParallelProcessingConfig()
    
    def excel_utility_router(self, cfg):
        """Route to appropriate task handler."""
        if cfg["task"] == "csv_copy_to_excel":
            self.csv_copy_to_excel_enhanced(cfg)
        return cfg
    
    def _process_single_csv_to_sheet(
        self, 
        csv_info: Dict[str, Any], 
        target_file: str,
        analysis_root_folder: str
    ) -> Tuple[bool, str]:
        """
        Process a single CSV file to an Excel sheet.
        
        Args:
            csv_info: Dictionary with 'input' and 'target' information
            target_file: Path to the target Excel file
            analysis_root_folder: Root folder for analysis
            
        Returns:
            Tuple of (success, message)
        """
        try:
            # Get input CSV file
            input_info = csv_info.get("input", {})
            input_csv = input_info.get("filename", "")
            
            if not input_csv:
                return False, "No input CSV specified"
            
            # Validate input file
            is_valid, input_csv = is_file_valid_func(input_csv, analysis_root_folder)
            if not is_valid:
                return False, f"Input file {input_csv} not found"
            
            # Get target sheet name
            target_info = csv_info.get("target", {})
            sheet_name = target_info.get("sheet_name", "Sheet1")
            
            # Read CSV data
            df = pd.read_csv(input_csv)
            
            # Load existing workbook
            wb = load_workbook(target_file, data_only=False)
            wb_sheetnames = wb.sheetnames
            
            if sheet_name not in wb_sheetnames:
                logger.info(f"Creating sheet {sheet_name} in {target_file}")
                wb.create_sheet(sheet_name)
            
            # Clear existing data in the sheet
            worksheet = wb[sheet_name]
            for row in worksheet["A1:BZ1000"]:
                for cell in row:
                    cell.value = None
            wb.save(target_file)
            wb.close()
            
            # Write CSV data to Excel
            with pd.ExcelWriter(
                target_file,
                mode="a",
                engine="openpyxl",
                if_sheet_exists="overlay",
            ) as writer:
                df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                    startrow=0,
                    startcol=0,
                )
            
            return True, f"Processed {os.path.basename(input_csv)} -> {sheet_name}"
            
        except Exception as e:
            error_msg = f"Error processing CSV: {str(e)}"
            logger.error(error_msg)
            return False, error_msg
    
    def csv_copy_to_excel_enhanced(self, cfg: Dict[str, Any]) -> Dict[str, Any]:
        """
        Enhanced CSV to Excel copying with parallel processing.
        Handles configurations with CSV lists inside groups.
        
        Args:
            cfg: Configuration dictionary
            
        Returns:
            Updated configuration with processing results
        """
        groups = cfg.get("data", {}).get("groups", [])
        
        if not groups:
            logger.warning("No groups found in configuration")
            return cfg
        
        # Check parallel processing configuration
        parallel_config = cfg.get("parallel_processing", {})
        enabled = parallel_config.get("enabled", True)
        
        # Track overall timing
        overall_start = time.time()
        
        all_results = []
        
        for group in groups:
            group_label = group.get("label", "unnamed")
            logger.info(f"\nProcessing group: {group_label}")
            
            # Get analysis root folder
            analysis_root_folder = cfg.get("Analysis", {}).get("analysis_root_folder", "")
            if not analysis_root_folder:
                analysis_root_folder = os.getcwd()
            
            # Get target file information
            target_info = group.get("target", {})
            template_file = target_info.get("template", "")
            target_file = target_info.get("filename", "")
            
            if not target_file:
                logger.error(f"No target filename specified for group {group_label}")
                continue
            
            # Validate target file path
            is_valid, target_file = is_file_valid_func(target_file, analysis_root_folder)
            if not is_valid:
                target_file = os.path.join(analysis_root_folder, target_file)
            
            # Create directory if needed
            target_dir = os.path.dirname(target_file)
            if target_dir and not os.path.exists(target_dir):
                os.makedirs(target_dir, exist_ok=True)
            
            # Copy template if specified
            if template_file:
                is_valid, template_file = is_file_valid_func(template_file, analysis_root_folder)
                if is_valid:
                    shutil.copy(template_file, target_file)
                    logger.info(f"Copied template to: {target_file}")
            
            # Create workbook if it doesn't exist
            if not os.path.exists(target_file):
                wb = Workbook(target_file)
                ws = wb.add_worksheet("Sheet1")
                wb.close()
                logger.info(f"Created new workbook: {target_file}")
            
            # Get CSV list
            csvs = group.get("csvs", [])
            
            if not csvs:
                logger.warning(f"No CSVs found for group {group_label}")
                continue
            
            logger.info(f"Found {len(csvs)} CSV files to process")
            
            # Process CSVs in parallel or sequential
            if not enabled or len(csvs) <= 1:
                # Sequential processing
                logger.info("Using sequential processing")
                group_start = time.time()
                
                for csv_info in csvs:
                    success, message = self._process_single_csv_to_sheet(
                        csv_info, target_file, analysis_root_folder
                    )
                    all_results.append({
                        'group': group_label,
                        'csv': csv_info,
                        'success': success,
                        'message': message
                    })
                    if success:
                        logger.info(f"  ✓ {message}")
                    else:
                        logger.warning(f"  ✗ {message}")
                
                group_time = time.time() - group_start
                logger.info(f"Group {group_label} completed in {group_time:.2f}s")
                
            else:
                # Parallel processing
                num_workers = self.parallel_config.get_optimal_workers(
                    num_tasks=len(csvs),
                    task_type=parallel_config.get("task_type", "io_bound")
                )
                
                logger.info(f"Using parallel processing with {num_workers} workers")
                group_start = time.time()
                
                with ThreadPoolExecutor(max_workers=num_workers) as executor:
                    # Submit all tasks
                    future_to_csv = {
                        executor.submit(
                            self._process_single_csv_to_sheet,
                            csv_info,
                            target_file,
                            analysis_root_folder
                        ): csv_info
                        for csv_info in csvs
                    }
                    
                    # Collect results as they complete
                    completed = 0
                    for future in as_completed(future_to_csv):
                        csv_info = future_to_csv[future]
                        try:
                            success, message = future.result()
                            completed += 1
                            all_results.append({
                                'group': group_label,
                                'csv': csv_info,
                                'success': success,
                                'message': message
                            })
                            
                            # Log progress
                            if success:
                                logger.info(f"  [{completed}/{len(csvs)}] ✓ {message}")
                            else:
                                logger.warning(f"  [{completed}/{len(csvs)}] ✗ {message}")
                                
                        except Exception as exc:
                            logger.error(f"CSV generated an exception: {exc}")
                            all_results.append({
                                'group': group_label,
                                'csv': csv_info,
                                'success': False,
                                'message': str(exc)
                            })
                
                group_time = time.time() - group_start
                logger.info(f"Group {group_label} completed in {group_time:.2f}s with {num_workers} workers")
        
        # Calculate overall statistics
        overall_time = time.time() - overall_start
        successful = sum(1 for r in all_results if r['success'])
        total = len(all_results)
        
        # Store results
        cfg['processing_results'] = all_results
        cfg['processing_stats'] = {
            'total_csvs': total,
            'successful': successful,
            'failed': total - successful,
            'total_time': overall_time,
            'avg_time_per_csv': overall_time / total if total > 0 else 0
        }
        
        # Log summary
        logger.info("\n" + "=" * 60)
        logger.info("PROCESSING COMPLETE")
        logger.info("=" * 60)
        logger.info(f"Total CSVs processed: {total}")
        logger.info(f"Successful: {successful}")
        logger.info(f"Failed: {total - successful}")
        logger.info(f"Total time: {overall_time:.2f}s")
        if total > 0:
            logger.info(f"Average time per CSV: {overall_time/total:.3f}s")
        logger.info("=" * 60)
        
        return cfg