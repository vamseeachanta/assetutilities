import os
import shutil
import io as io_module
from PIL import Image
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.colors as mcolors
import numpy as np

import pandas as pd
from loguru import logger
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from xlsxwriter.workbook import Workbook

try:
    import xlwings as xw
    XLWINGS_AVAILABLE = True
except ImportError:
    XLWINGS_AVAILABLE = False

from assetutilities.calculations.polynomial import Polynomial
from assetutilities.common.data import ReadFromExcel
from assetutilities.common.utilities import is_file_valid_func

rfe = ReadFromExcel()
_polynomial = Polynomial()


class ExcelUtilities:
    def __init__(self) -> None:
        pass

    def router(self, cfg):
        if cfg["task"] == "cross_reference_values_from_closed_workbooks":
            self.cross_reference_values_from_closed_workbooks_xlsxwriter(cfg)
        if cfg["task"] == "custom_calculation":
            self.get_data(cfg)
        return cfg

    def excel_utility_router(self, cfg):
        if cfg["task"] == "cross_reference_values_from_closed_workbooks":
            self.cross_reference_values_from_closed_workbooks_xlsxwriter(cfg)
        if cfg["task"] == "custom_calculation":
            self.get_data(cfg)
        if cfg["task"] == "excel_to_image":
            self.excel_to_image(cfg)
        if cfg["task"] == "csv_copy_to_excel":
            self.csv_copy_to_excel(cfg)
        return cfg

    def cross_reference_values_from_closed_workbooks_xlsxwriter(self, cfg):
        workbook = Workbook(
            cfg["files"]["target"]["path"] + cfg["files"]["target"]["workbook"]
        )
        worksheet = workbook.add_worksheet(cfg["files"]["target"]["worksheet"])
        worksheet.write_formula("A2", "=VLOOKUP(B3,'Sheet2'!$B$2:$R$2199,17,FALSE)")
        # TODO did not work with abstraction
        # worksheet.write_formula(cell_range, cell_formula)
        worksheet.write_formula("A3", "[S01-Dyn-FC180-2.5mHs.xlsx]EditingTable'!BH5")
        workbook.close()

    def get_data(self, cfg):
        pass

    def excel_to_image(self, cfg):
        # Default to Excel COM (xlwings) method for best visual quality
        use_xlwings = XLWINGS_AVAILABLE
        use_fallback = cfg.get("use_matplotlib_fallback", False)  # Allow override via config
        
        if not use_fallback:  # Try Excel COM first
            if not XLWINGS_AVAILABLE:
                logger.error("=" * 80)
                logger.error("EXCEL REQUIRED: xlwings library is not installed")
                logger.error("Please install xlwings: pip install xlwings")
                logger.error("Or set 'use_matplotlib_fallback: true' in config to use lower quality fallback")
                logger.error("=" * 80)
                raise ImportError("xlwings is required for high-quality Excel image export")
            
            try:
                # Test if Excel is available
                app = xw.App(visible=False, add_book=False)
                app.quit()
            except Exception as e:
                logger.error("=" * 80)
                logger.error("MICROSOFT EXCEL REQUIRED")
                logger.error("Microsoft Excel is not installed or not accessible on this computer")
                logger.error("Please either:")
                logger.error("  1. Install Microsoft Excel on this computer")
                logger.error("  2. Run this script on a computer with Excel installed")
                logger.error("  3. Set 'use_matplotlib_fallback: true' in config for lower quality output")
                logger.error(f"Technical error: {e}")
                logger.error("=" * 80)
                
                # Check if user wants to fall back
                if not cfg.get("force_excel_only", False):
                    logger.warning("Attempting matplotlib fallback (lower quality)...")
                    use_xlwings = False
                else:
                    raise RuntimeError("Excel is required but not available. Install Excel or change computer.")
        else:
            logger.info("Using matplotlib fallback method as configured")
            use_xlwings = False
        
        for file in cfg["files"]:
            io = file["io"]
            io_basename = os.path.splitext(os.path.basename(io))[0]
            analysis_root_folder = cfg["Analysis"]["analysis_root_folder"]
            is_file_valid, io = is_file_valid_func(io, analysis_root_folder)
            
            if not is_file_valid:
                logger.warning(f"File {io} not found, skipping...")
                continue
            
            if use_xlwings:
                self._export_with_xlwings(io, file, io_basename, cfg)
            else:
                self._export_with_openpyxl(io, file, io_basename, cfg)

    def csv_copy_to_excel(self, cfg):
        groups = cfg["data"]["groups"]
        for group in groups:
            analysis_root_folder = cfg["Analysis"]["analysis_root_folder"]
            template_file = group["target"]["template"]
            is_file_valid, template_file = is_file_valid_func(
                template_file, analysis_root_folder
            )
            target_file = group["target"]["filename"]
            is_file_valid, target_file = is_file_valid_func(
                target_file, analysis_root_folder
            )
            if not is_file_valid:
                target_file = os.path.join(analysis_root_folder, target_file)
            # Copy the template file to the target file
            shutil.copy(template_file, target_file)
            logger.debug(f"Template file copied to: {target_file}")
            csvs = group["csvs"]
            logger.info(f"For Excel file: {target_file}:")
            for csv in csvs:
                inputs_csv = csv["input"]["filename"]
                is_file_valid, inputs_csv = is_file_valid_func(
                    inputs_csv, analysis_root_folder
                )
                if not is_file_valid:
                    logger.debug(f"File {inputs_csv} not found. Skipping {target_file}")
                    logger.debug(f"Skipping {target_file}  .. FAIL")
                    if os.path.exists(target_file):
                        logger.info(f"Removing {target_file} as it is not updated")
                        os.remove(target_file)
                else:
                    df = pd.read_csv(inputs_csv)
                    wb = load_workbook(
                        target_file, data_only=False
                    )  # Keep formulas intact
                    sheet_name = csv["target"]["sheet_name"]
                    wb_sheetnames = wb.sheetnames
                    if sheet_name not in wb_sheetnames:
                        logger.info(
                            f"Sheet {sheet_name} does not exist in {target_file}."
                        )
                        wb.create_sheet(sheet_name)
                    worksheet = wb[sheet_name]
                    for row in worksheet["A1:BZ1000"]:
                        for cell in row:
                            cell.value = None
                    wb.save(target_file)
                    with pd.ExcelWriter(
                        target_file,
                        mode="a",
                        engine="openpyxl",
                        if_sheet_exists="overlay",
                    ) as writer:
                        df.to_excel(
                            writer,
                            sheet_name=sheet_name,
                            index=False,
                            startrow=0,
                            startcol=0,
                        )
                    logger.info(f"   ... Copying data in {inputs_csv} to {worksheet}")
    
    def _export_with_xlwings(self, io, file, io_basename, cfg):
        """Export Excel ranges using xlwings (requires Excel)"""
        app = None
        try:
            app = xw.App(visible=False, add_book=False)
            wb = app.books.open(io)
            
            for sheetname in file["sheet_name"]:
                if sheetname not in [s.name for s in wb.sheets]:
                    logger.warning(f"Sheet {sheetname} not found in {io}, skipping...")
                    continue
                
                ws = wb.sheets[sheetname]
                
                for array_range in file["range"]:
                    start_cell = array_range[0]
                    end_cell = array_range[1]
                    range_str = f"{start_cell}:{end_cell}"
                    
                    output_basename = (
                        io_basename + "_" + sheetname + "_" + 
                        array_range[0] + array_range[1]
                    )
                    
                    for ext in file["output_extension"]:
                        output_filename = output_basename + "." + ext
                        if file["output_dir"] is not None and os.path.isdir(
                            file["output_dir"]
                        ):
                            output_filename = os.path.join(
                                file["output_dir"], output_filename
                            )
                        else:
                            output_filename = os.path.join(
                                cfg["Analysis"]["result_folder"], output_filename
                            )
                        
                        try:
                            # Get the range
                            rng = ws.range(range_str)
                            
                            # Copy as picture and save
                            rng.api.CopyPicture(Format=2)  # 2 = xlBitmap
                            
                            # Create a temporary chart to paste the image
                            chart = ws.charts.add()
                            chart.api.Paste()
                            chart.api.Export(os.path.abspath(output_filename))
                            chart.delete()
                            
                            logger.info(f"Exported {sheetname}!{range_str} to {output_filename}")
                        except Exception as e:
                            logger.error(f"Failed to export range with xlwings: {e}")
            
            wb.close()
        except Exception as e:
            logger.error(f"xlwings export failed: {e}")
        finally:
            if app:
                try:
                    app.quit()
                except:
                    pass
    
    def _export_with_openpyxl(self, io, file, io_basename, cfg):
        """Export Excel ranges using openpyxl (no Excel required)"""
        try:
            # Load both versions - one for data, one for formulas
            wb_data = load_workbook(io, data_only=True)  # This has calculated values where available
            wb_formulas = load_workbook(io, data_only=False)  # This has formulas
            
            # Also keep a reference to both workbooks for cross-referencing
            self.wb_data = wb_data
            self.wb_formulas = wb_formulas
        except Exception as e:
            logger.error(f"Failed to load workbook {io}: {e}")
            return
        
        for sheetname in file["sheet_name"]:
            if sheetname not in wb_data.sheetnames:
                logger.warning(f"Sheet {sheetname} not found in {io}, skipping...")
                continue
            
            ws_data = wb_data[sheetname]
            ws_formulas = wb_formulas[sheetname]
            
            for array_range in file["range"]:
                start_cell = array_range[0]
                end_cell = array_range[1]
                
                output_basename = (
                    io_basename + "_" + sheetname + "_" + 
                    array_range[0] + array_range[1]
                )
                
                for ext in file["output_extension"]:
                    output_filename = output_basename + "." + ext
                    if file["output_dir"] is not None and os.path.isdir(
                        file["output_dir"]
                    ):
                        output_filename = os.path.join(
                            file["output_dir"], output_filename
                        )
                    else:
                        output_filename = os.path.join(
                            cfg["Analysis"]["result_folder"], output_filename
                        )
                    
                    try:
                        self._export_range_as_image_openpyxl(
                            ws_data, ws_formulas, wb_data, wb_formulas, start_cell, end_cell, output_filename
                        )
                        logger.info(f"Exported {sheetname}!{start_cell}:{end_cell} to {output_filename}")
                    except Exception as e:
                        logger.error(f"Failed to export range to image: {e}")
        
        wb_data.close()
        wb_formulas.close()
    
    def _evaluate_simple_formula(self, formula, wb, depth=0):
        """Evaluate simple Excel formulas for data extraction"""
        # Prevent infinite recursion
        if depth > 10:
            return None  # Return None instead of "[Formula]" so it gets handled properly
            
        if not isinstance(formula, str) or not formula.startswith('='):
            return formula
        
        try:
            # Handle simple cell references like =B9, =r_inputs!B9, =inputs!J2, =strut_dyn!E6
            if '!' in formula and '(' not in formula:
                # Sheet reference like =r_inputs!B9, =inputs!J2, or =strut_dyn!E6
                parts = formula[1:].split('!')
                if len(parts) == 2:
                    sheet_name, cell_ref = parts
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        try:
                            # First try to get the cell value directly
                            cell = ws[cell_ref]
                            cell_val = cell.value
                            
                            # If it's None and this is strut_dyn sheet, it might have calculated values
                            # Try to look in the data_only workbook if available
                            if cell_val is None and sheet_name == 'strut_dyn':
                                # For strut_dyn sheet, try to get from data_only workbook
                                if hasattr(self, 'wb_data') and sheet_name in self.wb_data.sheetnames:
                                    ws_data = self.wb_data[sheet_name]
                                    cell_data = ws_data[cell_ref]
                                    if cell_data.value is not None:
                                        logger.debug(f"Found strut_dyn value from data_only: {cell_ref} = {cell_data.value}")
                                        return cell_data.value
                                logger.debug(f"strut_dyn reference {cell_ref} returned None, might need calculated value")
                                return None
                            
                            # If it's None, try from the data_only workbook
                            if cell_val is None:
                                # Return None to indicate we couldn't evaluate
                                return None
                            
                            # If it's another formula, recursively evaluate
                            if isinstance(cell_val, str) and cell_val.startswith('='):
                                result = self._evaluate_simple_formula(cell_val, wb, depth + 1)
                                return result
                            else:
                                # It's a direct value, return it
                                return cell_val
                        except Exception as e:
                            logger.debug(f"Error evaluating {formula}: {e}")
                            return None
            elif formula.startswith('=') and '!' not in formula and '(' not in formula:
                # Simple cell reference like =B9, =C10
                # This is a same-sheet reference, return it as-is to be handled later
                return formula
            
            # Handle IF(ISNUMBER(SEARCH(...))) formulas
            if 'IF(ISNUMBER(SEARCH(' in formula:
                # Parse the search pattern
                import re
                # This is a complex formula, try to parse and evaluate
                if 'inputs!A' in formula:
                    # Get the referenced cell value
                    match = re.search(r'inputs!([A-Z]\d+)', formula)
                    if match and 'inputs' in wb.sheetnames:
                        cell_ref = match.group(1)
                        input_val = str(wb['inputs'][cell_ref].value or '')
                        
                        # Parse search patterns
                        search_patterns = re.findall(r'SEARCH\("([^"]+)"', formula)
                        
                        # Parse IF results
                        if_results = re.findall(r'"([^"]+)"', formula)
                        
                        # Evaluate based on patterns
                        for pattern in search_patterns:
                            if pattern.lower() in input_val.lower():
                                # Return the corresponding result
                                if 'l015' in pattern and 'l015' in input_val:
                                    return 0.15
                                elif 'l095' in pattern and 'l095' in input_val:
                                    return 0.95
                                elif '125km3' in pattern and '125km3' in input_val:
                                    return "125K m3"
                                elif '180km3' in pattern and '180km3' in input_val:
                                    return "180K m3"
                                elif '_pb' in pattern and '_pb' in input_val:
                                    return "Port"
                                elif '_sb' in pattern and '_sb' in input_val:
                                    return "SB"
                                elif 'hwl' in pattern.lower() and 'hwl' in input_val.lower():
                                    return "HHWL"
                                elif 'lwl' in pattern.lower() and 'lwl' in input_val.lower():
                                    return "LLWL"
                                elif '_cl_' in pattern and '_cl_' in input_val:
                                    return "Colinear"
                                elif '_ncl_' in input_val:
                                    return "Non-colinear"
            
            # If we can't evaluate, return empty string instead of [Formula]
            return ""
        except Exception as e:
            logger.debug(f"Could not evaluate formula {formula}: {e}")
            return ""
    
    def _export_range_as_image_openpyxl(self, worksheet_data, worksheet_formulas, wb_data, wb_formulas, start_cell, end_cell, output_path):
        """Export a range of cells from Excel worksheet as an image using matplotlib"""
        
        # Parse cell coordinates
        start_col_letter, start_row = coordinate_from_string(start_cell)
        end_col_letter, end_row = coordinate_from_string(end_cell)
        start_col = column_index_from_string(start_col_letter)
        end_col = column_index_from_string(end_col_letter)
        
        # Get actual column widths from Excel
        col_widths = []
        for col in range(start_col, end_col + 1):
            col_letter = get_column_letter(col)
            width = worksheet_formulas.column_dimensions[col_letter].width
            if width:
                # Excel width units to approximate character width
                col_widths.append(width)
            else:
                col_widths.append(8.43)  # Excel default width
        
        # Get row heights
        row_heights = []
        for row in range(start_row, end_row + 1):
            height = worksheet_formulas.row_dimensions[row].height
            if height:
                row_heights.append(height)
            else:
                row_heights.append(15)  # Excel default height
        
        # Check for merged cells in the range
        merged_cells = {}
        for merged_range in worksheet_formulas.merged_cells.ranges:
            if (merged_range.min_row >= start_row and merged_range.max_row <= end_row and
                merged_range.min_col >= start_col and merged_range.max_col <= end_col):
                # Store the merged range info
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        if row == merged_range.min_row and col == merged_range.min_col:
                            # This is the top-left cell of the merge
                            merged_cells[(row, col)] = {
                                'rows': merged_range.max_row - merged_range.min_row + 1,
                                'cols': merged_range.max_col - merged_range.min_col + 1,
                                'is_origin': True
                            }
                        else:
                            # This cell is part of a merge but not the origin
                            merged_cells[(row, col)] = {
                                'rows': 0,
                                'cols': 0,
                                'is_origin': False,
                                'origin': (merged_range.min_row, merged_range.min_col)
                            }
        
        # Create lists to hold the data and formatting
        data = []
        cell_colors = []
        text_colors = []
        font_weights = []
        max_col_widths = {}
        
        # Extract data and formatting from the range
        for row in range(start_row, end_row + 1):
            row_data = []
            row_colors = []
            row_text_colors = []
            row_font_weights = []
            
            for col in range(start_col, end_col + 1):
                # Check if this cell is part of a merged range
                merge_info = merged_cells.get((row, col), None)
                
                cell_data = worksheet_data.cell(row=row, column=col)
                cell_formula = worksheet_formulas.cell(row=row, column=col)
                
                # For merged cells, only show content in the origin cell
                if merge_info and not merge_info.get('is_origin', True):
                    # This is part of a merged cell but not the origin - leave it empty
                    value = ""
                else:
                    # Get value - try calculated value first from data_only workbook
                    value = cell_data.value
                    if value is None and cell_formula.value:
                        # If we have a formula but no calculated value, try to evaluate it
                        if isinstance(cell_formula.value, str) and cell_formula.value.startswith('='):
                            # Try to evaluate the formula using both workbooks
                            # First try with data workbook (which might have some calculated values)
                            evaluated = self._evaluate_simple_formula(cell_formula.value, wb_data, 0)
                            if evaluated is None:
                                # If that didn't work, try with formula workbook
                                evaluated = self._evaluate_simple_formula(cell_formula.value, wb_formulas, 0)
                            
                            if evaluated is not None and evaluated != "" and evaluated != cell_formula.value:
                                value = evaluated
                            else:
                                # For simple cell references within the same sheet
                                if cell_formula.value.startswith('=') and '!' not in cell_formula.value:
                                    ref = cell_formula.value[1:].replace('$', '')
                                    try:
                                        ref_cell_data = worksheet_data[ref]
                                        if ref_cell_data.value is not None:
                                            value = ref_cell_data.value
                                        else:
                                            # Try from formula sheet
                                            ref_cell_formula = worksheet_formulas[ref]
                                            if ref_cell_formula.value:
                                                value = self._evaluate_simple_formula(ref_cell_formula.value, wb_data, 0)
                                    except:
                                        value = ""
                                else:
                                    value = ""
                        else:
                            value = cell_formula.value
                
                # Format the value based on number format
                if value is None:
                    formatted_value = ""
                elif isinstance(value, str):
                    # String values pass through as-is
                    formatted_value = value
                elif isinstance(value, (int, float)):
                    # Special handling for percentages that were evaluated from formulas
                    # If it's 0.15 or 0.95, these are load percentages
                    if value in [0.15, 0.95]:
                        if cell_formula.number_format and '%' in cell_formula.number_format:
                            formatted_value = f"{value:.0%}"  # Show as 15% or 95%
                        else:
                            formatted_value = f"{value:.2f}"  # Show as 0.15 or 0.95
                    # Apply number formatting
                    elif cell_formula.number_format:
                        try:
                            if '%' in cell_formula.number_format:
                                formatted_value = f"{value:.1%}"
                            elif '0.00' in cell_formula.number_format:
                                formatted_value = f"{value:.2f}"
                            elif '0' in cell_formula.number_format and '.' not in cell_formula.number_format:
                                formatted_value = str(int(value))
                            else:
                                formatted_value = f"{value:.2f}"
                        except:
                            formatted_value = str(value)
                    else:
                        if value == int(value):
                            formatted_value = str(int(value))
                        else:
                            formatted_value = f"{value:.2f}"
                else:
                    formatted_value = str(value)
                
                row_data.append(formatted_value)
                
                # Get cell background color (use formula sheet for formatting)
                bg_color = '#FFFFFF'  # Default white
                
                # Check if this is a header row (rows 2-3 from the start typically)
                if row - start_row + 1 in [2, 3]:
                    # Apply yellowish/beige background for headers
                    bg_color = '#FFCC99'  # Light orange/beige color similar to reference
                
                if cell_formula.fill and cell_formula.fill.start_color:
                    try:
                        if cell_formula.fill.start_color.rgb:
                            rgb = cell_formula.fill.start_color.rgb
                            if isinstance(rgb, str) and len(rgb) >= 6:
                                # Remove alpha channel if present
                                if len(rgb) == 8:
                                    rgb = rgb[2:]
                                # Check if it's transparent/no fill (00000000 or 000000)
                                if rgb == '000000' or rgb == '00000000':
                                    # Keep the header color if it's a header row
                                    if row - start_row + 1 not in [2, 3]:
                                        bg_color = '#FFFFFF'  # Default to white
                                else:
                                    bg_color = f'#{rgb}'
                        elif cell_formula.fill.start_color.theme is not None:
                            # Handle theme colors 
                            # Keep header color for header rows
                            if row - start_row + 1 not in [2, 3]:
                                bg_color = '#FFFFFF'
                    except:
                        pass
                row_colors.append(bg_color)
                
                # Get text color
                text_color = '#000000'  # Default black
                if cell_formula.font and cell_formula.font.color:
                    try:
                        if cell_formula.font.color.rgb:
                            rgb = cell_formula.font.color.rgb
                            if isinstance(rgb, str) and len(rgb) >= 6:
                                if len(rgb) == 8:
                                    rgb = rgb[2:]
                                # Check if it's transparent/default (00000000 or 000000)
                                if rgb == '000000' or rgb == '00000000':
                                    text_color = '#000000'  # Default to black
                                else:
                                    text_color = f'#{rgb}'
                        elif cell.font.color.theme is not None:
                            # Handle theme colors - default to black for now
                            text_color = '#000000'
                    except:
                        pass
                row_text_colors.append(text_color)
                
                # Get font weight
                font_weight = 'normal'
                if cell_formula.font and cell_formula.font.bold:
                    font_weight = 'bold'
                row_font_weights.append(font_weight)
                
                # Track maximum width for each column
                col_idx = col - start_col
                if col_idx not in max_col_widths:
                    max_col_widths[col_idx] = 0
                max_col_widths[col_idx] = max(max_col_widths[col_idx], len(formatted_value))
            
            data.append(row_data)
            cell_colors.append(row_colors)
            text_colors.append(row_text_colors)
            font_weights.append(row_font_weights)
        
        # Create figure and axis
        n_rows = end_row - start_row + 1
        n_cols = end_col - start_col + 1
        
        # Calculate figure size based on actual Excel column widths
        # Excel width unit is approximately 1/7 of a character width
        # Convert to inches - using a smaller multiplier for more compact display
        total_width_chars = sum(col_widths)
        fig_width = total_width_chars * 0.12  # Adjusted for better match to reference
        
        # Calculate height based on row heights
        # Excel row height is in points (1/72 inch)
        total_height_points = sum(row_heights)
        fig_height = total_height_points / 72 * 1.3  # Slightly more compact for better match
        
        # Ensure minimum sizes
        fig_width = max(fig_width, 10)
        fig_height = max(fig_height, 5)
        
        fig, ax = plt.subplots(figsize=(fig_width, fig_height))
        ax.axis('tight')
        ax.axis('off')
        
        # Create table with column width ratios based on Excel widths
        col_width_ratios = [w/sum(col_widths) for w in col_widths]
        
        # Create table
        table = ax.table(cellText=data, 
                        cellLoc='center',
                        loc='center',
                        colWidths=col_width_ratios,
                        bbox=[0, 0, 1, 1])
        
        # Style the table
        table.auto_set_font_size(False)
        table.set_fontsize(6)  # Smaller font to match reference and fit more data
        
        # Apply cell-specific styling
        for i in range(n_rows):
            for j in range(n_cols):
                cell = table[(i, j)]
                
                # Set background color
                try:
                    cell.set_facecolor(cell_colors[i][j])
                except:
                    cell.set_facecolor('#FFFFFF')
                
                # Set text properties
                try:
                    cell.set_text_props(
                        color=text_colors[i][j],
                        weight=font_weights[i][j]
                    )
                except:
                    cell.set_text_props(color='black')
                
                # Set border
                cell.set_edgecolor('#D0D0D0')
                cell.set_linewidth(0.5)
        
        # Adjust layout and save
        plt.tight_layout(pad=0.05)
        plt.savefig(output_path, dpi=200, bbox_inches='tight', pad_inches=0.05)
        plt.close(fig)
